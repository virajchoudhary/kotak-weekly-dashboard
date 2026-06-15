import io
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))

import weekly_engine


class WeeklyEngineTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.original_db = weekly_engine.DB_PATH
        cls.temp_dir = tempfile.TemporaryDirectory()
        weekly_engine.DB_PATH = Path(cls.temp_dir.name) / "weekly_test.db"
        cls.fixture = Path(__file__).resolve().parents[1] / "data" / "uploads" / "Weekly Inflow And AUM-1st-31st May'26.xlsx"
        cls.upload_bytes = cls.fixture.read_bytes()
        weekly_engine.init_db()

    @classmethod
    def tearDownClass(cls):
        weekly_engine.DB_PATH = cls.original_db
        cls.temp_dir.cleanup()

    def test_period_detection_from_summary_title(self):
        period = weekly_engine.parse_period(
            "Weekly Inflow And AUM-1st-31st May'26.xlsx",
            "Summary-  01ST To 31ST MAY'26",
        )
        self.assertEqual(period["periodKey"], "2026-05-31")
        self.assertEqual(period["periodLabel"], "01ST To 31ST MAY'26")
        self.assertEqual(period["financialYear"], "2026-2027")

    def test_row_projection_contains_screenshot_labels(self):
        period, rows, schemes = weekly_engine.parse_upload(self.upload_bytes, self.fixture.name)
        labels = [row["rowLabel"] for row in rows if row["inSummary"]]
        self.assertEqual(labels[0], "ARBITRAGE FUND Total")
        self.assertIn("Dividend Yield Fund", labels)
        self.assertIn("Value Fund/Contra Fund", labels)
        self.assertEqual(labels[-3:], ["Liquid Total", "DEBT", "Total"])
        self.assertGreater(len(schemes), 100)

    def test_summary_workbook_shape_and_formatting(self):
        weekly_engine.process_upload(self.upload_bytes, self.fixture.name)
        output = weekly_engine.compile_summary_workbook(period_key="2026-05-31")
        wb = load_workbook(io.BytesIO(output), data_only=False)
        ws = wb["Summary"]
        self.assertEqual(ws.max_row, 29)
        self.assertEqual(ws.max_column, 10)
        self.assertEqual(ws["A1"].value, "Summary- 01ST To 31ST MAY'26")
        self.assertEqual(ws["A4"].value, "ARBITRAGE FUND Total")
        self.assertEqual(ws["A28"].value, "DEBT")
        self.assertEqual(ws["A29"].value, "Total")
        self.assertEqual(ws.column_dimensions["A"].width, weekly_engine.SPEC_COLUMN_WIDTH_A)
        self.assertEqual(ws.column_dimensions["B"].width, weekly_engine.SPEC_COLUMN_WIDTH_DATA)
        self.assertEqual(ws.row_dimensions[1].height, weekly_engine.ROW_HEIGHT)
        self.assertEqual(ws.row_dimensions[29].height, weekly_engine.ROW_HEIGHT)
        self.assertEqual(ws["D4"].number_format, weekly_engine.PERCENT_FORMAT)
        self.assertEqual(ws["A5"].fill.fgColor.rgb[-6:], weekly_engine.YELLOW)
        self.assertEqual(ws["A29"].fill.fgColor.rgb[-6:], weekly_engine.TOTAL_BLUE)

    def test_same_month_upload_replaces_and_mom_has_ytd_summary_block(self):
        period, rows, schemes = weekly_engine.parse_upload(self.upload_bytes, self.fixture.name)
        conn = weekly_engine.get_db_connection()
        try:
            weekly_engine.upsert_upload(conn, period, rows, schemes, self.fixture.name)
            weekly_engine.upsert_upload(conn, period, rows, schemes, self.fixture.name)
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM periods WHERE period_key = ?", (period["periodKey"],))
            self.assertEqual(cur.fetchone()[0], 1)

            period2 = {
                **period,
                "periodKey": "2026-06-07",
                "periodLabel": "01ST To 07TH JUN'26",
                "startDate": "2026-06-01",
                "endDate": "2026-06-07",
            }
            weekly_engine.upsert_upload(conn, period2, rows, schemes, "Weekly Inflow And AUM-1st-7th Jun'26.xlsx")
            period3 = {
                **period,
                "periodKey": "2026-06-14",
                "periodLabel": "01ST To 14TH JUN'26",
                "startDate": "2026-06-01",
                "endDate": "2026-06-14",
            }
            weekly_engine.upsert_upload(conn, period3, rows, schemes, "Weekly Inflow And AUM-1st-14th Jun'26.xlsx")
            cur.execute("SELECT COUNT(*) FROM periods WHERE financial_year = ?", (period["financialYear"],))
            self.assertEqual(cur.fetchone()[0], 2)
            cur.execute("SELECT period_key FROM periods WHERE substr(end_date, 1, 7) = '2026-06'")
            self.assertEqual(cur.fetchone()[0], "2026-06-14")
        finally:
            conn.close()

        output = weekly_engine.compile_mom_workbook(fy="2026-2027")
        wb = load_workbook(io.BytesIO(output), data_only=False)
        ws = wb["MoM-YTD"]
        self.assertEqual(ws.max_row, 41)
        self.assertEqual(ws.max_column, 28)
        self.assertEqual(ws["B1"].value, "Summary-  (01ST To 31ST MAY'26)")
        self.assertEqual(ws["B2"].value, "AUM (as on 31th MAY'26)")
        self.assertEqual(ws["E2"].value, "Gross Sales")
        self.assertEqual(ws["H2"].value, "Net Sales")
        self.assertEqual(ws["K1"].value, "Summary-  (01ST To 14TH JUN'26)")
        self.assertEqual(ws["K2"].value, "AUM (as on 14th JUN'26)")
        self.assertEqual(ws["N2"].value, "Gross Sales")
        self.assertEqual(ws["Q2"].value, "Net Sales")
        self.assertEqual(ws["T1"].value, "Summary-  (01ST MAY To 14TH JUN'26)")
        self.assertEqual(ws["T2"].value, "AUM (as on 14th JUN'26)")
        self.assertEqual(ws["W2"].value, "YTD Gross Sales")
        self.assertEqual(ws["Z2"].value, "YTD Net Sales")
        self.assertEqual(ws.freeze_panes, "B4")
        self.assertEqual(ws.column_dimensions["A"].width, weekly_engine.SPEC_COLUMN_WIDTH_A)
        self.assertEqual(ws.column_dimensions["B"].width, weekly_engine.SPEC_COLUMN_WIDTH_DATA)
        self.assertEqual(ws.row_dimensions[41].height, weekly_engine.ROW_HEIGHT)
        total_row = 41
        self.assertEqual(ws.cell(total_row, 20).value, "=K41")
        self.assertEqual(ws.cell(total_row, 23).value, "=SUM(E41,N41)")
        self.assertEqual(ws.cell(total_row, 26).value, "=SUM(H41,Q41)")


if __name__ == "__main__":
    unittest.main()
