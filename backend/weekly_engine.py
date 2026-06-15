import calendar
import io
import re
import sqlite3
from copy import copy
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.views import Selection

DB_PATH = Path(__file__).resolve().parent / "weekly_amfi.db"
PROJECT_ROOT = Path(__file__).resolve().parents[1]
SEED_UPLOAD_DIRS = (
    PROJECT_ROOT / "data" / "uploads",
    PROJECT_ROOT / "templates",
    PROJECT_ROOT,
)

SUMMARY_ROWS = (
    7, 9, 11, 13, 16,
    17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28,
    29, 31, 34, 36, 37, 38, 41, 80, 81,
)

MOM_ROWS = (
    7, 9, 11, 13, 16,
    17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28,
    29, 31, 34, 36, 37, 38,
    41, 45, 47, 49, 52, 54, 61, 63, 66, 68, 70, 75, 79, 80, 81,
)

YELLOW_ROWS = {
    "Equity Savings Total", "Balanced Total", "ELSS Total",
    "Large & Mid Cap Fund", "Multi Cap Fund", "Multi Asset Allocation",
    "Value Fund/Contra Fund",
}

BLUE_ROWS = {"Mid Cap Fund", "Sectoral/Thematic Funds"}
PEACH_ROWS = {
    "Dividend Yield Fund", "Equity Others (C)", "Focused Fund", "Large Cap Fund",
    "Small Cap Fund", "Flexi Cap Fund", "Equity Total", "INDEX FUND Total",
}
GRAY_ROWS = {
    "ARBITRAGE FUND Total", "BAAF Total", "Solution Oriented Total", "OTHER ETF Total",
    "PURE EQUITY", "Liquid Total", "DEBT ETF Total", "FMP Total", "FMP",
    "CAPITAL PROTECT Total", "CAPITAL PROTECT", "Credit Risk Total",
    "Debt Hybrid Total", "Duration Total", "Other Debt Total", "FOF-OVERSEAS Total",
    "Others Total", "Short Duration Total", "ULTRA SHORT TERM Total",
}

MONTHS = {
    "jan": 1, "january": 1, "feb": 2, "february": 2, "mar": 3, "march": 3,
    "apr": 4, "april": 4, "may": 5, "jun": 6, "june": 6, "jul": 7, "july": 7,
    "aug": 8, "august": 8, "sep": 9, "sept": 9, "september": 9, "oct": 10,
    "october": 10, "nov": 11, "november": 11, "dec": 12, "december": 12,
}
MONTH_ABBR = {v: k.title() for k, v in {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6, "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}.items()}

HEADER_FILL = "DDEBF7"
TOTAL_BLUE = "BDD7EE"
SECTION_ORANGE = "F8CBAD"
LIGHT_PEACH = "FCE4D6"
LIGHT_BLUE = "D9E2F3"
GRAY = "D9D9D9"
YELLOW = "FFFF00"
WHITE = "FFFFFF"
BLACK = "000000"

NUMBER_FORMAT = '#,##0;(#,##0);-'
PERCENT_FORMAT = '0.00%'
NEG_PERCENT_FORMAT = '0.00%;-0.00%;0.00%'
ROW_HEIGHT = 14.5
SPEC_COLUMN_WIDTH_A = 26.0
SPEC_COLUMN_WIDTH_DATA = 12.2
WORKBOOK_TEMPLATE_VERSION = "monthly-blocks-with-final-ytd-summary-v2"


def get_db_connection():
    conn = sqlite3.connect(DB_PATH, timeout=10.0)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    return conn


def init_db():
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS periods (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                period_key TEXT NOT NULL UNIQUE,
                period_label TEXT NOT NULL,
                start_date TEXT NOT NULL,
                end_date TEXT NOT NULL,
                financial_year TEXT NOT NULL,
                source_filename TEXT,
                uploaded_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS summary_rows (
                period_key TEXT NOT NULL,
                source_row INTEGER NOT NULL,
                row_order INTEGER NOT NULL,
                row_label TEXT NOT NULL,
                in_summary INTEGER NOT NULL,
                in_mom INTEGER NOT NULL,
                aum_kotak REAL,
                aum_industry REAL,
                aum_ms,
                gross_kotak REAL,
                gross_industry REAL,
                gross_ms,
                net_kotak REAL,
                net_industry REAL,
                net_ms,
                PRIMARY KEY (period_key, source_row),
                FOREIGN KEY (period_key) REFERENCES periods(period_key) ON DELETE CASCADE
            )
        """)
        cur.execute("""
            CREATE TABLE IF NOT EXISTS scheme_rows (
                period_key TEXT NOT NULL,
                scheme_name TEXT NOT NULL,
                asset_class TEXT,
                asset_amc TEXT,
                sales_mis_group TEXT,
                scheme_main_group TEXT,
                aum REAL,
                gross_sales REAL,
                net_sales REAL,
                redemption REAL,
                PRIMARY KEY (period_key, scheme_name),
                FOREIGN KEY (period_key) REFERENCES periods(period_key) ON DELETE CASCADE
            )
        """)
        cur.execute("UPDATE summary_rows SET row_order = source_row WHERE row_order != source_row")
        conn.commit()
        seed_initial_upload_if_empty(conn)
        collapse_monthly_uploads(conn)
    finally:
        conn.close()


def seed_initial_upload_if_empty(conn):
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM periods")
    if cur.fetchone()[0]:
        return
    candidates = []
    for upload_dir in SEED_UPLOAD_DIRS:
        if upload_dir.exists():
            candidates.extend(sorted(upload_dir.glob("*.xlsx")))
    for candidate in candidates:
        try:
            period, rows, schemes = parse_upload(candidate.read_bytes(), candidate.name)
        except ValueError:
            continue
        upsert_upload(conn, period, rows, schemes, candidate.name)
        return


def collapse_monthly_uploads(conn):
    cur = conn.cursor()
    duplicate_months = cur.execute("""
        SELECT financial_year, substr(end_date, 1, 7) AS month_key
        FROM periods
        GROUP BY financial_year, substr(end_date, 1, 7)
        HAVING COUNT(*) > 1
    """).fetchall()
    for row in duplicate_months:
        keep = cur.execute("""
            SELECT period_key
            FROM periods
            WHERE financial_year = ?
              AND substr(end_date, 1, 7) = ?
            ORDER BY end_date DESC, uploaded_at DESC, period_key DESC
            LIMIT 1
        """, (row["financial_year"], row["month_key"])).fetchone()
        if not keep:
            continue
        obsolete_keys = [
            item["period_key"]
            for item in cur.execute("""
                SELECT period_key
                FROM periods
                WHERE financial_year = ?
                  AND substr(end_date, 1, 7) = ?
                  AND period_key != ?
            """, (row["financial_year"], row["month_key"], keep["period_key"])).fetchall()
        ]
        for key in obsolete_keys:
            cur.execute("DELETE FROM summary_rows WHERE period_key = ?", (key,))
            cur.execute("DELETE FROM scheme_rows WHERE period_key = ?", (key,))
            cur.execute("DELETE FROM periods WHERE period_key = ?", (key,))
    conn.commit()


def parse_upload(upload_bytes: bytes, filename: str) -> tuple[dict, list[dict], list[dict]]:
    workbook_values = load_workbook(io.BytesIO(upload_bytes), data_only=True)
    workbook_formulas = load_workbook(io.BytesIO(upload_bytes), data_only=False)
    try:
        if "Summary" not in workbook_values.sheetnames:
            raise ValueError("Uploaded workbook must contain a 'Summary' sheet.")
        ws_values = workbook_values["Summary"]
        ws_formulas = workbook_formulas["Summary"]
        title = str(ws_formulas["C3"].value or ws_values["C3"].value or "")
        period = parse_period(filename, title)
        rows = extract_summary_rows(ws_values)
        schemes = extract_scheme_rows(workbook_values)
        return period, rows, schemes
    finally:
        workbook_values.close()
        workbook_formulas.close()


def parse_period(filename: str, title: str) -> dict:
    text = f"{title} {filename}"
    match = re.search(
        r"(?P<start>\d{1,2})(?:ST|ND|RD|TH|st|nd|rd|th)?\s*(?:to|-)\s*"
        r"(?P<end>\d{1,2})(?:ST|ND|RD|TH|st|nd|rd|th)?\s*"
        r"(?P<month>[A-Za-z]{3,9})'?[-\s]*(?P<year>\d{2,4})",
        text,
        flags=re.IGNORECASE,
    )
    if not match:
        raise ValueError("Could not detect weekly period from Summary title or filename.")
    month_raw = match.group("month").lower()
    month = MONTHS.get(month_raw[:3]) or MONTHS.get(month_raw)
    if not month:
        raise ValueError(f"Unknown month in weekly period: {month_raw}")
    year_raw = match.group("year")
    year = int(year_raw) if len(year_raw) == 4 else 2000 + int(year_raw)
    start_day = int(match.group("start"))
    end_day = int(match.group("end"))
    last_day = calendar.monthrange(year, month)[1]
    if not (1 <= start_day <= last_day and 1 <= end_day <= last_day):
        raise ValueError("Detected weekly period has an invalid day.")
    start = date(year, month, start_day)
    end = date(year, month, end_day)
    if end < start:
        raise ValueError("Detected weekly period end date is before start date.")
    period_label = f"{ordinal(start_day)} To {ordinal(end_day)} {MONTH_ABBR[month].upper()}'{str(year)[-2:]}"
    period_key = end.isoformat()
    fy = f"{year}-{year + 1}" if month >= 4 else f"{year - 1}-{year}"
    return {
        "periodKey": period_key,
        "periodLabel": period_label,
        "startDate": start.isoformat(),
        "endDate": end.isoformat(),
        "financialYear": fy,
    }


def ordinal(day: int) -> str:
    if 10 <= day % 100 <= 20:
        suffix = "TH"
    else:
        suffix = {1: "ST", 2: "ND", 3: "RD"}.get(day % 10, "TH")
    return f"{day:02d}{suffix}"


def extract_summary_rows(ws) -> list[dict]:
    output = []
    all_rows = tuple(dict.fromkeys((*SUMMARY_ROWS, *MOM_ROWS)))
    for row_idx in all_rows:
        label = display_label(ws, row_idx)
        if not label:
            continue
        output.append({
            "sourceRow": row_idx,
            "rowOrder": row_idx,
            "rowLabel": label,
            "inSummary": 1 if row_idx in SUMMARY_ROWS else 0,
            "inMom": 1 if row_idx in MOM_ROWS else 0,
            "aumKotak": clean_value(ws.cell(row_idx, 3).value),
            "aumIndustry": clean_value(ws.cell(row_idx, 4).value),
            "aumMs": clean_value(ws.cell(row_idx, 5).value),
            "grossKotak": clean_value(ws.cell(row_idx, 6).value),
            "grossIndustry": clean_value(ws.cell(row_idx, 7).value),
            "grossMs": clean_value(ws.cell(row_idx, 8).value),
            "netKotak": clean_value(ws.cell(row_idx, 9).value),
            "netIndustry": clean_value(ws.cell(row_idx, 10).value),
            "netMs": clean_value(ws.cell(row_idx, 11).value),
        })
    return output


def display_label(ws, row_idx: int) -> str:
    label = ws.cell(row_idx, 1).value
    asset_class = ws.cell(row_idx, 2).value
    if row_idx in range(17, 29):
        label = asset_class
    text = str(label or "").strip()
    if text == "FMP Total":
        return "FMP"
    if text == "CAPITAL PROTECT Total":
        return "CAPITAL PROTECT"
    return text


def extract_scheme_rows(wb) -> list[dict]:
    if "Scheme Wise" not in wb.sheetnames:
        return []
    ws = wb["Scheme Wise"]
    rows = []
    for row_idx in range(2, ws.max_row + 1):
        scheme_name = str(ws.cell(row_idx, 1).value or "").strip()
        if not scheme_name or is_scheme_aggregate_row(scheme_name):
            continue
        rows.append({
            "schemeName": scheme_name,
            "assetClass": ws.cell(row_idx, 2).value,
            "assetAmc": ws.cell(row_idx, 3).value,
            "salesMisGroup": ws.cell(row_idx, 4).value,
            "schemeMainGroup": ws.cell(row_idx, 5).value,
            "aum": clean_value(ws.cell(row_idx, 6).value),
            "grossSales": clean_value(ws.cell(row_idx, 7).value),
            "netSales": clean_value(ws.cell(row_idx, 8).value),
            "redemption": clean_value(ws.cell(row_idx, 9).value),
        })
    return rows


def is_scheme_aggregate_row(value: str) -> bool:
    key = re.sub(r"[^a-z0-9]+", " ", str(value or "").strip().lower()).strip()
    return key in {"grand total", "total"} or key.endswith(" total")


def clean_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        if text.upper() in {"N.A", "NA", "N/A"}:
            return "N.A"
        try:
            return float(text.replace(",", ""))
        except ValueError:
            return text
    return float(value) if isinstance(value, int | float) else value


def process_upload(upload_bytes: bytes, filename: str) -> tuple[dict, list[str]]:
    period, rows, schemes = parse_upload(upload_bytes, filename)
    conn = get_db_connection()
    try:
        upsert_upload(conn, period, rows, schemes, filename)
    finally:
        conn.close()
    return period, []


def upsert_upload(conn, period: dict, rows: list[dict], schemes: list[dict], filename: str | None = None):
    cur = conn.cursor()
    period_key = period["periodKey"]
    month_prefix = period["endDate"][:7]
    existing_keys = [
        row[0]
        for row in cur.execute(
            """
            SELECT period_key
            FROM periods
            WHERE financial_year = ?
              AND substr(end_date, 1, 7) = ?
            """,
            (period["financialYear"], month_prefix),
        ).fetchall()
    ]
    if period_key not in existing_keys:
        existing_keys.append(period_key)
    for key in existing_keys:
        cur.execute("DELETE FROM summary_rows WHERE period_key = ?", (key,))
        cur.execute("DELETE FROM scheme_rows WHERE period_key = ?", (key,))
        cur.execute("DELETE FROM periods WHERE period_key = ?", (key,))
    cur.execute("""
        INSERT INTO periods (
            period_key, period_label, start_date, end_date, financial_year, source_filename
        ) VALUES (?, ?, ?, ?, ?, ?)
    """, (
        period_key, period["periodLabel"], period["startDate"], period["endDate"],
        period["financialYear"], filename,
    ))
    cur.executemany("""
        INSERT INTO summary_rows (
            period_key, source_row, row_order, row_label, in_summary, in_mom,
            aum_kotak, aum_industry, aum_ms, gross_kotak, gross_industry,
            gross_ms, net_kotak, net_industry, net_ms
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, [
        (
            period_key, row["sourceRow"], row["rowOrder"], row["rowLabel"],
            row["inSummary"], row["inMom"], row["aumKotak"], row["aumIndustry"], row["aumMs"],
            row["grossKotak"], row["grossIndustry"], row["grossMs"],
            row["netKotak"], row["netIndustry"], row["netMs"],
        )
        for row in rows
    ])
    cur.executemany("""
        INSERT INTO scheme_rows (
            period_key, scheme_name, asset_class, asset_amc, sales_mis_group,
            scheme_main_group, aum, gross_sales, net_sales, redemption
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, [
        (
            period_key, row["schemeName"], row["assetClass"], row["assetAmc"],
            row["salesMisGroup"], row["schemeMainGroup"], row["aum"],
            row["grossSales"], row["netSales"], row["redemption"],
        )
        for row in schemes
    ])
    conn.commit()


def list_archives() -> list[dict]:
    conn = get_db_connection()
    try:
        cur = conn.cursor()
        rows = cur.execute("""
            SELECT financial_year, COUNT(*) AS period_count, MAX(uploaded_at) AS last_modified
            FROM periods
            GROUP BY financial_year
            ORDER BY financial_year DESC
        """).fetchall()
        return [
            {
                "financial_year": row["financial_year"],
                "period_count": row["period_count"],
                "last_modified": row["last_modified"],
                "status": "Ready",
            }
            for row in rows
        ]
    finally:
        conn.close()


def latest_financial_year() -> str | None:
    conn = get_db_connection()
    try:
        row = conn.execute("SELECT financial_year FROM periods ORDER BY end_date DESC LIMIT 1").fetchone()
        return row["financial_year"] if row else None
    finally:
        conn.close()


def periods_for_fy(fy: str) -> list[dict]:
    conn = get_db_connection()
    try:
        rows = conn.execute("""
            SELECT period_key, period_label, start_date, end_date, financial_year, uploaded_at
            FROM periods
            WHERE financial_year = ?
            ORDER BY end_date
        """, (fy,)).fetchall()
        return [dict(row) for row in rows]
    finally:
        conn.close()


def rows_for_period(period_key: str, summary_only: bool = False, mom_only: bool = False) -> list[dict]:
    predicate = ""
    if summary_only:
        predicate = "AND in_summary = 1"
    if mom_only:
        predicate = "AND in_mom = 1"
    conn = get_db_connection()
    try:
        rows = conn.execute(f"""
            SELECT *
            FROM summary_rows
            WHERE period_key = ? {predicate}
            ORDER BY row_order
        """, (period_key,)).fetchall()
        return [dict(row) for row in rows]
    finally:
        conn.close()


def dashboard_payload(
    fy: str | None = None,
    period_key: str | None = None,
    warnings: list[str] | None = None,
    upload_period: dict | None = None,
) -> dict:
    target_fy = fy or latest_financial_year()
    if not target_fy:
        return {"summary": {}, "timeSeries": [], "categorySummary": [], "schemeSummary": [], "warnings": warnings or [], "financialYear": None}
    periods = periods_for_fy(target_fy)
    if not periods:
        return {"summary": {}, "timeSeries": [], "categorySummary": [], "schemeSummary": [], "warnings": warnings or [], "financialYear": target_fy}
    selected = next((period for period in periods if period["period_key"] == period_key), periods[-1])
    selected_rows = rows_for_period(selected["period_key"], mom_only=True)
    total_row = next((row for row in selected_rows if row["row_label"] == "Total"), {})
    equity_row = next((row for row in selected_rows if row["row_label"] == "Equity"), {})
    debt_row = next((row for row in selected_rows if row["row_label"] == "DEBT"), {})
    time_series = []
    for period in periods:
        row = next((item for item in rows_for_period(period["period_key"], mom_only=True) if item["row_label"] == "Total"), {})
        time_series.append({
            "periodKey": period["period_key"],
            "period": period["period_label"],
            "periodShort": short_period_label(period),
            "endDate": period["end_date"],
            "aum": round_num(row.get("aum_kotak")),
            "grossSales": round_num(row.get("gross_kotak")),
            "netSales": round_num(row.get("net_kotak")),
        })
    category_summary = build_category_summary(selected_rows)
    scheme_summary = build_scheme_summary(selected["period_key"])
    return {
        "financialYear": target_fy,
        "uploadPeriod": upload_period,
        "latestPeriod": periods[-1]["period_label"],
        "selectedPeriodKey": selected["period_key"],
        "selectedPeriod": selected["period_label"],
        "selectedPeriodShort": short_period_label(selected),
        "displayRange": display_range_label(periods),
        "periods": periods,
        "summary": {
            "latestPeriod": selected["period_label"],
            "latestAum": round_num(total_row.get("aum_kotak")),
            "latestIndustryAum": round_num(total_row.get("aum_industry")),
            "latestGrossSales": round_num(total_row.get("gross_kotak")),
            "latestNetSales": round_num(total_row.get("net_kotak")),
            "equityAum": round_num(equity_row.get("aum_kotak")),
            "debtAum": round_num(debt_row.get("aum_kotak")),
        },
        "timeSeries": time_series,
        "categorySummary": category_summary,
        "schemeSummary": scheme_summary,
        "warnings": warnings or [],
        "generatedAt": datetime.utcnow().isoformat(timespec="seconds") + "Z",
    }


def build_category_summary(rows: list[dict]) -> list[dict]:
    labels = ["ARBITRAGE FUND Total", "Equity Savings Total", "PURE EQUITY", "Equity", "Liquid Total", "DEBT"]
    output = []
    total_aum = next((row.get("aum_kotak") for row in rows if row["row_label"] == "Total"), 0) or 0
    for row in rows:
        if row["row_label"] not in labels:
            continue
        aum = float(row.get("aum_kotak") or 0)
        output.append({
            "category": row["row_label"],
            "latestAum": round_num(aum),
            "grossSales": round_num(row.get("gross_kotak")),
            "netSales": round_num(row.get("net_kotak")),
            "marketShare": percent_or_none(row.get("aum_ms")),
            "aumShare": 0 if not total_aum else round(aum / total_aum, 6),
        })
    return output


def short_period_label(period: dict) -> str:
    start = datetime.fromisoformat(period["start_date"]).date()
    end = datetime.fromisoformat(period["end_date"]).date()
    return f"{start.day:02d}-{end.day:02d} {MONTH_ABBR[end.month]}'{str(end.year)[-2:]}"


def display_range_label(periods: list[dict]) -> str:
    if not periods:
        return "-"
    first = datetime.fromisoformat(periods[0]["start_date"]).date()
    last = datetime.fromisoformat(periods[-1]["end_date"]).date()
    return f"{first.day:02d} {MONTH_ABBR[first.month]}'{str(first.year)[-2:]} to {last.day:02d} {MONTH_ABBR[last.month]}'{str(last.year)[-2:]}"


def build_scheme_summary(period_key: str) -> list[dict]:
    conn = get_db_connection()
    try:
        rows = conn.execute("""
            SELECT *
            FROM scheme_rows
            WHERE period_key = ?
              AND lower(trim(scheme_name)) NOT IN ('grand total', 'total')
            ORDER BY aum DESC
        """, (period_key,)).fetchall()
        return [
            {
                "schemeKey": row["scheme_name"],
                "schemeName": row["scheme_name"],
                "assetClass": row["asset_class"],
                "assetAmc": row["asset_amc"],
                "mainGroup": row["scheme_main_group"],
                "aum": round_num(row["aum"]),
                "grossSales": round_num(row["gross_sales"]),
                "netSales": round_num(row["net_sales"]),
                "redemption": round_num(row["redemption"]),
            }
            for row in rows
        ]
    finally:
        conn.close()


def compile_summary_workbook(period_key: str | None = None, fy: str | None = None) -> bytes:
    period = resolve_period(period_key, fy)
    rows = rows_for_period(period["period_key"], summary_only=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    setup_summary_sheet(ws, f"Summary- {period['period_label']}")
    for idx, row in enumerate(rows, start=4):
        write_report_row(ws, idx, row, summary=True)
    apply_outer_borders(ws, 1, 1, 29, 10)
    finalize_sheet(wb, ws)
    return save_workbook(wb)


def compile_mom_workbook(fy: str | None = None) -> bytes:
    target_fy = fy or latest_financial_year()
    if not target_fy:
        raise ValueError("No weekly uploads found.")
    periods = periods_for_fy(target_fy)
    if not periods:
        raise ValueError(f"No weekly uploads found for FY {target_fy}.")
    wb = Workbook()
    ws = wb.active
    ws.title = "MoM-YTD"
    ws.cell(3, 1).value = "ASSET (AMC)"
    style_header_cell(ws.cell(3, 1))
    all_rows = [
        {"source_row": row["source_row"], "row_label": row["row_label"]}
        for row in rows_for_period(periods[-1]["period_key"], mom_only=True)
    ]
    for block_idx, period in enumerate(periods):
        rows = rows_for_period(period["period_key"], mom_only=True)
        row_map = {row["source_row"]: row for row in rows}
        start_col = 2 + block_idx * 9
        setup_mom_block(ws, start_col, period)
        for row_pos, identity in enumerate(all_rows, start=4):
            source = row_map.get(identity["source_row"], {})
            label = identity["row_label"]
            row_data = {
                "row_label": label,
                "aum_kotak": source.get("aum_kotak"),
                "aum_industry": source.get("aum_industry"),
                "aum_ms": ratio(source.get("aum_kotak"), source.get("aum_industry")),
                "gross_kotak": source.get("gross_kotak"),
                "gross_industry": source.get("gross_industry"),
                "gross_ms": ratio(source.get("gross_kotak"), source.get("gross_industry")),
                "net_kotak": source.get("net_kotak"),
                "net_industry": source.get("net_industry"),
                "net_ms": ratio(source.get("net_kotak"), source.get("net_industry"), allow_negative=True),
            }
            if block_idx == 0:
                ws.cell(row_pos, 1).value = label
                style_row_cell(ws.cell(row_pos, 1), label, summary=False)
            write_mom_row(ws, row_pos, start_col, row_data)
    ytd_start_col = 2 + len(periods) * 9
    setup_ytd_summary_block(ws, ytd_start_col, periods)
    latest_start_col = ytd_start_col - 9
    gross_cols = [2 + idx * 9 + 3 for idx in range(len(periods))]
    net_cols = [2 + idx * 9 + 6 for idx in range(len(periods))]
    for row_pos, identity in enumerate(all_rows, start=4):
        write_ytd_summary_row(
            ws,
            row_pos,
            ytd_start_col,
            identity["row_label"],
            latest_start_col,
            gross_cols,
            net_cols,
        )
    end_col = 1 + (len(periods) + 1) * 9
    ws.freeze_panes = "B4"
    ws.column_dimensions["A"].width = SPEC_COLUMN_WIDTH_A
    for col in range(2, end_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = SPEC_COLUMN_WIDTH_DATA
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = ROW_HEIGHT
    apply_outer_borders(ws, 1, 1, ws.max_row, end_col)
    finalize_sheet(wb, ws)
    return save_workbook(wb)


def resolve_period(period_key: str | None = None, fy: str | None = None) -> dict:
    conn = get_db_connection()
    try:
        if period_key:
            row = conn.execute("SELECT * FROM periods WHERE period_key = ?", (period_key,)).fetchone()
        elif fy:
            row = conn.execute("SELECT * FROM periods WHERE financial_year = ? ORDER BY end_date DESC LIMIT 1", (fy,)).fetchone()
        else:
            row = conn.execute("SELECT * FROM periods ORDER BY end_date DESC LIMIT 1").fetchone()
        if not row:
            raise ValueError("No weekly upload found for the requested period.")
        return dict(row)
    finally:
        conn.close()


def setup_summary_sheet(ws, title: str):
    ws.merge_cells("A1:J1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", bold=True, size=11, color=BLACK)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("B2:D2")
    ws.merge_cells("E2:G2")
    ws.merge_cells("H2:J2")
    for cell, value in (("B2", "AUM"), ("E2", "Gross Sales"), ("H2", "Net Sales")):
        ws[cell] = value
        style_header_cell(ws[cell], size=11)
    headers = ["ASSET (AMC)", "Kotak", "Industry", "MS", "Kotak", "Industry", "MS", "Kotak", "Industry", "MS"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(3, col)
        cell.value = header
        style_header_cell(cell, size=11)
    ws.column_dimensions["A"].width = SPEC_COLUMN_WIDTH_A
    for col in range(2, 11):
        ws.column_dimensions[get_column_letter(col)].width = SPEC_COLUMN_WIDTH_DATA
    for row in range(1, 30):
        ws.row_dimensions[row].height = ROW_HEIGHT


def setup_mom_block(ws, start_col: int, period: dict):
    period_end = datetime.fromisoformat(period["end_date"]).date()
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 8)
    title_cell = ws.cell(1, start_col)
    title_cell.value = f"Summary-  ({period['period_label']})"
    title_cell.font = Font(name="Calibri", bold=True, size=11, color=BLACK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = thin_border()
    ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=start_col + 2)
    ws.merge_cells(start_row=2, start_column=start_col + 3, end_row=2, end_column=start_col + 5)
    ws.merge_cells(start_row=2, start_column=start_col + 6, end_row=2, end_column=start_col + 8)
    headings = [
        (start_col, f"AUM (as on {period_end.day:02d}th {MONTH_ABBR[period_end.month].upper()}'{str(period_end.year)[-2:]})"),
        (start_col + 3, "Gross Sales"),
        (start_col + 6, "Net Sales"),
    ]
    for col, text in headings:
        cell = ws.cell(2, col)
        cell.value = text
        style_header_cell(cell)
    for offset, header in enumerate(["Kotak", "Industry", "MS"] * 3):
        cell = ws.cell(3, start_col + offset)
        cell.value = header
        style_header_cell(cell)


def setup_ytd_summary_block(ws, start_col: int, periods: list[dict]):
    latest = periods[-1]
    latest_end = datetime.fromisoformat(latest["end_date"]).date()
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 8)
    title_cell = ws.cell(1, start_col)
    title_cell.value = ytd_summary_title(periods)
    title_cell.font = Font(name="Calibri", bold=True, size=11, color=BLACK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = thin_border()
    ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=start_col + 2)
    ws.merge_cells(start_row=2, start_column=start_col + 3, end_row=2, end_column=start_col + 5)
    ws.merge_cells(start_row=2, start_column=start_col + 6, end_row=2, end_column=start_col + 8)
    headings = [
        (start_col, f"AUM (as on {latest_end.day:02d}th {MONTH_ABBR[latest_end.month].upper()}'{str(latest_end.year)[-2:]})"),
        (start_col + 3, "YTD Gross Sales"),
        (start_col + 6, "YTD Net Sales"),
    ]
    for col, text in headings:
        cell = ws.cell(2, col)
        cell.value = text
        style_header_cell(cell)
    for offset, header in enumerate(["Kotak", "Industry", "MS"] * 3):
        cell = ws.cell(3, start_col + offset)
        cell.value = header
        style_header_cell(cell)


def ytd_summary_title(periods: list[dict]) -> str:
    first_start = datetime.fromisoformat(periods[0]["start_date"]).date()
    latest_end = datetime.fromisoformat(periods[-1]["end_date"]).date()
    start_text = f"{ordinal(first_start.day)} {MONTH_ABBR[first_start.month].upper()}"
    end_text = f"{ordinal(latest_end.day)} {MONTH_ABBR[latest_end.month].upper()}'{str(latest_end.year)[-2:]}"
    return f"Summary-  ({start_text} To {end_text})"


def write_report_row(ws, target_row: int, row: dict, summary: bool):
    ws.cell(target_row, 1).value = row["row_label"]
    values = [
        row["aum_kotak"], row["aum_industry"], row["aum_ms"],
        row["gross_kotak"], row["gross_industry"], row["gross_ms"],
        row["net_kotak"], row["net_industry"], row["net_ms"],
    ]
    for col, value in enumerate(values, start=2):
        ws.cell(target_row, col).value = value
    for col in range(1, 11):
        style_row_cell(ws.cell(target_row, col), row["row_label"], summary=summary)
        if col in {4, 7, 10}:
            ws.cell(target_row, col).number_format = PERCENT_FORMAT
        elif col > 1:
            ws.cell(target_row, col).number_format = NUMBER_FORMAT
    ws.row_dimensions[target_row].height = ROW_HEIGHT


def write_mom_row(ws, target_row: int, start_col: int, row: dict):
    values = [
        row["aum_kotak"], row["aum_industry"], row["aum_ms"],
        row["gross_kotak"], row["gross_industry"], row["gross_ms"],
        row["net_kotak"], row["net_industry"], row["net_ms"],
    ]
    for offset, value in enumerate(values):
        cell = ws.cell(target_row, start_col + offset)
        cell.value = value
        style_row_cell(cell, row["row_label"], summary=False)
        if offset in {2, 5, 8}:
            cell.number_format = NEG_PERCENT_FORMAT if offset == 8 else PERCENT_FORMAT
        else:
            cell.number_format = NUMBER_FORMAT


def write_ytd_summary_row(
    ws,
    target_row: int,
    start_col: int,
    label: str,
    latest_start_col: int,
    gross_cols: list[int],
    net_cols: list[int],
):
    latest_aum_kotak = ws.cell(target_row, latest_start_col).coordinate
    latest_aum_industry = ws.cell(target_row, latest_start_col + 1).coordinate
    aum_kotak = ws.cell(target_row, start_col)
    aum_industry = ws.cell(target_row, start_col + 1)
    aum_ms = ws.cell(target_row, start_col + 2)
    aum_kotak.value = f"={latest_aum_kotak}"
    aum_industry.value = f"={latest_aum_industry}"
    aum_ms.value = f"=IFERROR({aum_kotak.coordinate}/{aum_industry.coordinate},0)"

    gross_kotak = ws.cell(target_row, start_col + 3)
    gross_industry = ws.cell(target_row, start_col + 4)
    gross_ms = ws.cell(target_row, start_col + 5)
    gross_kotak.value = sum_formula(ws, target_row, gross_cols)
    gross_industry.value = sum_formula(ws, target_row, [col + 1 for col in gross_cols])
    gross_ms.value = f"=IFERROR({gross_kotak.coordinate}/{gross_industry.coordinate},0)"

    net_kotak = ws.cell(target_row, start_col + 6)
    net_industry = ws.cell(target_row, start_col + 7)
    net_ms = ws.cell(target_row, start_col + 8)
    net_kotak.value = sum_formula(ws, target_row, net_cols)
    net_industry.value = sum_formula(ws, target_row, [col + 1 for col in net_cols])
    net_ms.value = f'=IF(OR({net_kotak.coordinate}<=0,{net_industry.coordinate}<=0),"N.A",{net_kotak.coordinate}/{net_industry.coordinate})'

    for offset in range(9):
        cell = ws.cell(target_row, start_col + offset)
        style_row_cell(cell, label, summary=False)
        if offset in {2, 5, 8}:
            cell.number_format = NEG_PERCENT_FORMAT if offset == 8 else PERCENT_FORMAT
        else:
            cell.number_format = NUMBER_FORMAT


def sum_formula(ws, target_row: int, cols: list[int]) -> str:
    refs = ",".join(ws.cell(target_row, col).coordinate for col in cols)
    return f"=SUM({refs})"


def style_header_cell(cell, size: int = 11):
    cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
    cell.font = Font(name="Calibri", bold=True, color=BLACK, size=size)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()


def style_row_cell(cell, label: str, summary: bool = False):
    fill = row_fill(label, summary=summary)
    cell.fill = PatternFill("solid", fgColor=fill)
    bold = is_bold_row(label)
    cell.font = Font(name="Calibri", bold=bold, color=BLACK, size=10 if fill != TOTAL_BLUE else 11)
    cell.alignment = Alignment(horizontal="left" if cell.column == 1 else "center", vertical="center")
    cell.border = thin_border()


def row_fill(label: str, summary: bool = False) -> str:
    if label == "Total":
        return TOTAL_BLUE
    if summary and label in {"PURE EQUITY", "Equity", "DEBT"}:
        return SECTION_ORANGE
    if label in {"Equity", "DEBT"}:
        return SECTION_ORANGE
    if label in YELLOW_ROWS:
        return YELLOW
    if label in BLUE_ROWS:
        return LIGHT_BLUE
    if label in PEACH_ROWS:
        return LIGHT_PEACH
    if label in GRAY_ROWS:
        return GRAY
    return LIGHT_PEACH


def is_bold_row(label: str) -> bool:
    return (
        label.endswith("Total")
        or label in {"FMP", "CAPITAL PROTECT", "PURE EQUITY", "Equity", "DEBT", "Total"}
        or label in YELLOW_ROWS
        or label in BLUE_ROWS
    )


def thin_border():
    side = Side(style="thin", color=BLACK)
    return Border(left=side, right=side, top=side, bottom=side)


def apply_outer_borders(ws, min_row: int, min_col: int, max_row: int, max_col: int):
    border = thin_border()
    for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = copy(border)


def finalize_sheet(wb, ws):
    ws.sheet_view.showGridLines = False
    ws.sheet_view.topLeftCell = "A1"
    if ws.sheet_view.selection:
        ws.sheet_view.selection[0].activeCell = "A1"
        ws.sheet_view.selection[0].sqref = "A1"
    else:
        ws.sheet_view.selection = [Selection(activeCell="A1", sqref="A1")]
    wb.active = wb.index(ws)


def save_workbook(wb) -> bytes:
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def ratio(numerator, denominator, allow_negative: bool = False):
    if numerator is None or denominator in (None, 0):
        return "N.A" if allow_negative else 0
    if allow_negative and (float(numerator) <= 0 or float(denominator) <= 0):
        return "N.A"
    try:
        return float(numerator) / float(denominator)
    except ZeroDivisionError:
        return "N.A" if allow_negative else 0


def round_num(value):
    if value is None or isinstance(value, str):
        return value
    return round(float(value), 2)


def percent_or_none(value):
    if value is None or isinstance(value, str):
        return None
    return round(float(value), 6)
